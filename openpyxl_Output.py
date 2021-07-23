# -*- coding: utf-8 -*-
"""
Created on Sun Jun 20 11:23:31 2021

@author: wb580236
"""
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.chart import BarChart, Series, Reference
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
#from Transp_GHG_Tool_v1 import TransGHG
import pdb 

class openpyxl_Output():
    
    def __init__(self,filename):
        self.filename = filename
        
    def detailOutput(self,cntry_df, filename ):
        self.cntry_df = cntry_df
        self.filename = filename
        wb = Workbook()
        ##### First Sheet to add the detailed data
        ws1 = wb.create_sheet("Detailed")
        for r in dataframe_to_rows(cntry_df, index=True, header=True):
            ws1.append(r)
        wb.save(filename)
    
    def sheetOutput(self, cntry_df,cntry_agg_df, cntry_yr_agg_df, region_df, cntry, filename):
        self.cntry_df = cntry_df
        self.cntry_agg_df = cntry_agg_df
        self.cntry_yr_agg_df = cntry_yr_agg_df
        self.region_df = region_df
        #avg_reg_df = region_df.groupby('Mode').mean()
        self.cntry = cntry
        self.filename  = filename
        wb = Workbook()
        ##### First Sheet to add the detailed data
        ws1 = wb.create_sheet("Detailed")
        for r in dataframe_to_rows(cntry_df, index=True, header=True):
            ws1.append(r)
        
        
        
        #### Second Sheet with aggregated country data and 
        cntry_agg_df_sub = cntry_agg_df[["Mode", "Tot_GHG"]]
        ws2 = wb.create_sheet("Comparison")
        for r in dataframe_to_rows(cntry_agg_df_sub, index=True, header=True):
            ws2.append(r)
        
        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.grouping = "percentStacked"
        #chart1.overlap = 100
        chart1.overlap = 100
        chart1.title = cntry + ' Contribution from Transport Modes'
        chart1.y_axis.title = 'Percentage of CO2eq GHG Emissions'
        #chart1.x_axis.title = 'x-axis'
        
        data = Reference(ws2, min_col=2, min_row=3, max_row=10, max_col=3)
        cats = Reference(ws2, min_col=3, min_row=1, max_col=3)
        chart1.add_data(data, from_rows=True, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.shape = 4
        ws2.add_chart(chart1, "K2")
        
        
        region_df_mean = region_df.groupby('Mode').mean()
        region_df_mean.reset_index(inplace = True)
        region_df_sub = region_df_mean[["Mode", "Tot_GHG"]]
        
        region_list = region_df_sub.values.tolist()
        ws2.cell(row=14, column=2).value = 'Mode'
        ws2.cell(row=14, column=3).value = 'Tot_HG'
        for rowN in range(0,len(region_list)):
            for colN in range(0,len(region_list[rowN])):
                val = region_list[rowN][colN]
                ws2.cell(row=rowN+15, column=colN+2, value=val)
                #"Comparison".cell(row=rowN+10,column=colN).value = val
             
        chart11 = BarChart()
        chart11.type = "col"
        chart11.style = 10
        chart11.grouping = "percentStacked"
        chart11.overlap = 100
        #chart11.gap = 150
        chart11.title = 'Regional Contribution from Transport Modes'
        chart11.y_axis.title = 'Percentage of CO2eq GHG Emissions'
        #chart1.x_axis.title = 'x-axis'
        
        data = Reference(ws2, min_col=2, min_row=15, max_row=22, max_col=3)
        cats = Reference(ws2, min_col=3, min_row=14, max_col=3)
        chart11.add_data(data, from_rows=True, titles_from_data=True)
        chart11.set_categories(cats)
        chart11.shape = 4
        ws2.add_chart(chart11, "K20")
        
         
        #### Third Sheet with          
        ws3 = wb.create_sheet("Scatter_Chart")
        for r in dataframe_to_rows(cntry_yr_agg_df, index=True, header=True):
            ws3.append(r)
        
 
        chart2 = ScatterChart()
      
        for i in range(3, 67,8):
            
            seriesname = ws3.cell(row=i, column=3).value
            xvalues = Reference(ws3, min_col=4, min_row=i, max_row= i + 7)
            values = Reference(ws3, min_col=8, min_row=i, max_row=i + 7)
            series = Series(values, xvalues, title=seriesname)
            chart2.series.append(series)
        
        chart2.title = "Time Series of Transportation Emissions"
        chart2.style = 13
        chart2.x_axis.title = 'Years'
        chart2.y_axis.title = 'GHG Emissions in kTon CO2eq'
        chart2.legend.position = 'r'

        ws3.add_chart(chart2, "K20")
        
        wb.save(filename)
        #pdb.set_trace()
 
#def main():
#    cntry = 'India'
#    filename = "C:\\Users\\wb580236\\OneDrive - WBG\\TransportDecarbonization_Lit\\charts_test.xlsx"
#    newTransGHG  = TransGHG('SQLite')
#    newTransGHG.FuelCalc()
#    newTransGHG.genERtabl('Default')
#    newTransGHG.EmisCalc()
#    cntryAgg_df = newTransGHG.qryAggData(cntry,'Aggregated')
#    cntryDet_df = newTransGHG.qryAggData(cntry,'Detailed')
#    cntryYrly_df = newTransGHG.qryAggData(cntry,'YearlyAgg')
#    simAgg_df = pd.read_csv("C:\\Users\\wb580236\\OneDrive - WBG\\TransportDecarbonization_Lit\\test.csv")
#    newOutput =  openpyxl_Output(filename)
#    newOutput.sheetOutput(cntryDet_df,cntryAgg_df,cntryYrly_df,simAgg_df,cntry,filename )
            
#if __name__ == "__main__":
#    main()